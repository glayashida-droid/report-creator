"""Resolve OEM/customer-specific report rules from report_templates/special.xlsx."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from src.io.network_sources import report_templates_directory
from src.language_copy import has_chinese
from src.models.project_state import ProjectState, SpecialReportProfile

_SPECIAL_FILENAME = "special.xlsx"
_YES = frozenset({"是", "yes", "y", "true", "1"})

# Default lab footer addresses (templates use {{实验室地址_*}} placeholders).
DEFAULT_LAB_ADDRESS_CN = "上海市闵行区万芳路1351号"
DEFAULT_LAB_ADDRESS_EN = (
    "Certre Testing International Pinzheng (Shanghai)Co., Ltd.        "
    "No.1351, Wanfang Road, Minhang, Shanghai, China"
)

# Legacy marker strings (keyword fallback removed; kept for tests/docs).
_DEFAULT_LAB_CN = (
    DEFAULT_LAB_ADDRESS_CN,
    "上海市闵行区新骏环路777号5号楼",
)
_DEFAULT_LAB_EN = (
    DEFAULT_LAB_ADDRESS_EN,
    "Centre Testing International Pinzheng(Shanghai) Co., Ltd.  Building 5, No. 777, Xinjun Ring Road, Minhang District, Shanghai, China",
)


@dataclass(frozen=True)
class SpecialRuleRow:
    oem_cn: str = ""
    oem_alias: str = ""
    client_cn: str = ""
    client_alias: str = ""
    show_test_period: bool = False
    show_tester: bool = False
    use_4sign: bool = False
    forbid_na: bool = False
    lab_address: str = ""


def _normalize(text: str) -> str:
    text = (text or "").strip()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    return text.lower() if not has_chinese(text) else text


def rule_keyword_matches(keyword: str, application_value: str) -> bool:
    """Short name in special.xlsx matches application value by containment."""
    kw = _normalize(keyword)
    val = _normalize(application_value)
    if not kw or not val:
        return False
    if has_chinese(kw):
        return kw in val or val in kw
    kl = kw.lower()
    raw = (application_value or "").strip()
    vl = val.lower()
    if kl in vl or vl in kl:
        return True
    from difflib import SequenceMatcher

    for token in re.findall(r"[A-Za-z]+", raw):
        tl = token.lower()
        if kl in tl or tl in kl:
            return True
        if len(kl) >= 4 and len(tl) >= 4 and SequenceMatcher(None, kl, tl).ratio() >= 0.84:
            return True
    return False


def _is_yes(value: object) -> bool:
    return str(value or "").strip().lower() in _YES


def _split_lab_address(raw: str) -> Tuple[str, str]:
    text = (raw or "").strip()
    if not text:
        return "", ""
    if "/" in text:
        left, right = text.split("/", 1)
        cn = left.strip()
        en = right.strip()
        if has_chinese(cn):
            return cn, en
        if has_chinese(right.strip()):
            return right.strip(), left.strip()
        return cn, en
    if has_chinese(text):
        return text, ""
    return "", text


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_sheet_rows(ws) -> List[SpecialRuleRow]:
    rows: List[SpecialRuleRow] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(_cell_str(c) for c in row):
            continue
        cells = list(row) + [None] * max(0, 9 - len(row))
        oem_cn = _cell_str(cells[0])
        oem_alias = _cell_str(cells[1])
        client_cn = _cell_str(cells[2])
        client_alias = _cell_str(cells[3])
        if not any((oem_cn, oem_alias, client_cn, client_alias)):
            # Address-only rows (e.g. 洛柯) still participate in address resolution.
            address_only = _cell_str(cells[8])
            if address_only and oem_cn and not _is_yes(_cell_str(cells[4])):
                rows.append(
                    SpecialRuleRow(
                        oem_cn=oem_cn,
                        lab_address=address_only,
                    )
                )
            continue
        rows.append(
            SpecialRuleRow(
                oem_cn=oem_cn,
                oem_alias=oem_alias,
                client_cn=client_cn,
                client_alias=client_alias,
                show_test_period=_is_yes(cells[4]),
                show_tester=_is_yes(cells[5]),
                use_4sign=_is_yes(cells[6]),
                forbid_na=_is_yes(cells[7]),
                lab_address=_cell_str(cells[8]),
            )
        )
    return rows


def load_special_rules(path: Optional[Path] = None) -> List[SpecialRuleRow]:
    if path is None:
        folder = report_templates_directory()
        path = folder / _SPECIAL_FILENAME if folder else None
    if path is None or not path.is_file():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        return _parse_sheet_rows(ws)
    finally:
        wb.close()


def _row_matches_oem(row: SpecialRuleRow, oem_values: Sequence[str]) -> bool:
    keys = [row.oem_cn, row.oem_alias]
    for key in keys:
        if not key:
            continue
        for val in oem_values:
            if rule_keyword_matches(key, val):
                return True
    return False


def _row_matches_client(row: SpecialRuleRow, client_values: Sequence[str]) -> bool:
    keys = [row.client_cn, row.client_alias]
    for key in keys:
        if not key:
            continue
        for val in client_values:
            if rule_keyword_matches(key, val):
                return True
    return False


def _oem_values_from_state(state: ProjectState) -> List[str]:
    fields = state.application_fields or {}
    fields_en = state.application_fields_en or {}
    vals = [
        fields.get("主机厂", ""),
        fields_en.get("主机厂", ""),
    ]
    return [v for v in vals if (v or "").strip()]


def _client_values_from_state(state: ProjectState) -> List[str]:
    cn = (state.report_title_name or state.applicant_name or "").strip()
    en = (state.report_title_name_en or state.applicant_name_en or "").strip()
    fields = state.application_fields or {}
    fields_en = state.application_fields_en or {}
    vals = [
        cn,
        en,
        fields.get("申请公司", ""),
        fields_en.get("申请公司", ""),
    ]
    seen = set()
    out: List[str] = []
    for v in vals:
        v = (v or "").strip()
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


def resolve_special_profile(
    state: ProjectState,
    rules: Optional[Sequence[SpecialRuleRow]] = None,
    rules_path: Optional[Path] = None,
) -> SpecialReportProfile:
    """Match OEM/customer rows; boolean flags union; lab address OEM-first."""
    if rules is None:
        rules = load_special_rules(rules_path)

    oem_vals = _oem_values_from_state(state)
    client_vals = _client_values_from_state(state)

    show_test_period = False
    show_tester = False
    use_4sign = False
    forbid_na = False
    oem_address = ""
    client_address = ""

    for row in rules:
        oem_hit = _row_matches_oem(row, oem_vals)
        client_hit = _row_matches_client(row, client_vals)
        if not oem_hit and not client_hit:
            continue
        if oem_hit or client_hit:
            show_test_period = show_test_period or row.show_test_period
            show_tester = show_tester or row.show_tester
            use_4sign = use_4sign or row.use_4sign
            forbid_na = forbid_na or row.forbid_na
        if oem_hit and row.lab_address:
            oem_address = row.lab_address
        elif client_hit and row.lab_address and not oem_address:
            client_address = row.lab_address

    address_raw = oem_address or client_address
    cn, en = _split_lab_address(address_raw)
    return SpecialReportProfile(
        show_test_period=show_test_period,
        show_tester=show_tester,
        use_4sign=use_4sign,
        forbid_na=forbid_na,
        lab_address_cn=cn,
        lab_address_en=en,
    )

def refresh_special_profile(
    state: ProjectState,
    rules_path: Optional[Path] = None,
) -> SpecialReportProfile:
    profile = resolve_special_profile(state, rules_path=rules_path)
    state.special_profile = profile
    return profile


def profile_from_state(state: ProjectState) -> SpecialReportProfile:
    raw = getattr(state, "special_profile", None)
    if isinstance(raw, SpecialReportProfile):
        return raw
    if isinstance(raw, dict):
        return SpecialReportProfile(**raw)
    return SpecialReportProfile()


def state_has_forbidden_na(state: ProjectState, leg_filter: Optional[str] = None) -> List[str]:
    """Return labels of tests that still have N/A when forbid_na is active."""
    from src.models.project_state import TestResult

    profile = profile_from_state(state)
    if not profile.forbid_na:
        return []
    labels: List[str] = []
    for leg, node in state.iter_nodes_for_export(leg_filter):
        for sample in node.samples or []:
            results = sample.all_results()
            if any(r == TestResult.NA for r in results):
                name = (node.test_name or "").strip() or "（未命名试验）"
                label = f"{leg.leg_name} / {name}"
                if label not in labels:
                    labels.append(label)
                break
    return labels


def default_lab_address_markers() -> Tuple[Tuple[str, ...], Tuple[str, ...]]:
    return _DEFAULT_LAB_CN, _DEFAULT_LAB_EN
