"""Manage data-table xlsx attachments under 3.测试组/{Leg名}-{试验名}/数据表附件/."""

from __future__ import annotations

import re
import shutil
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries

from application_parser.sample_id_labels import is_sample_id_column_key

from src.language_copy import table_header_label
from src.io.network_sources import data_table_templates_directory
from src.io.test_photos import (
    PhotoError,
    TEST_GROUP_DIR,
    require_leg_name as require_photo_leg_name,
    require_usable_test_name as require_photo_test_name,
    test_dir,
)
from src.models.project_state import DataTableRef, TestNode

ATTACHMENT_DIR = "数据表附件"
_BAD_NAME = re.compile(r'[\\/:*?"<>|]')

# macOS .app bundle names / Windows executable stems, preferred order
_EXCEL_APP_NAMES = ("Microsoft Excel",)
_WPS_APP_NAMES = ("wpsoffice", "WPS Office", "kingsoft")


def default_templates_dir() -> Path:
    return data_table_templates_directory()


class DataTableError(Exception):
    pass


def rewrite_test_dir_in_relative_path(
    relative_path: str, old_dir_key: str, new_dir_key: str
) -> str:
    """Rewrite 3.测试组/{old}/… → 3.测试组/{new}/… in a stored attachment path."""
    old = (old_dir_key or "").strip()
    new = (new_dir_key or "").strip()
    rel = (relative_path or "").replace("\\", "/")
    if not old or not new or old == new or not rel:
        return relative_path or ""
    old_prefix = f"{TEST_GROUP_DIR}/{old}/"
    if rel.startswith(old_prefix):
        return f"{TEST_GROUP_DIR}/{new}/" + rel[len(old_prefix) :]
    return rel


def retarget_node_data_tables(node: TestNode, old_dir_key: str, new_dir_key: str) -> None:
    """Update data_tables relative_path prefixes after a trial folder rename."""
    refs = list(getattr(node, "data_tables", None) or [])
    if not refs:
        return
    node.data_tables = [
        DataTableRef(
            title=ref.title,
            relative_path=rewrite_test_dir_in_relative_path(
                ref.relative_path, old_dir_key, new_dir_key
            ),
        )
        for ref in refs
    ]


@dataclass
class PreviewSnapshot:
    """Readonly first-sheet bbox: values as strings, empty cells kept, merges as A1:B2."""

    sheet_name: str
    values: List[List[str]] = field(default_factory=list)
    merges: List[str] = field(default_factory=list)
    # 1-based origin of values[0][0] in the sheet (for mapping merge refs onto the grid)
    origin_row: int = 1
    origin_col: int = 1


def attachment_dir(project_root: Path, leg_name: str, test_name: str) -> Path:
    return test_dir(project_root, leg_name, test_name) / ATTACHMENT_DIR


def _require_leg_name(leg_name: str) -> str:
    try:
        return require_photo_leg_name(leg_name)
    except PhotoError as exc:
        raise DataTableError(str(exc)) from exc


def _require_usable_test_name(test_name: str) -> str:
    try:
        return require_photo_test_name(test_name)
    except PhotoError as exc:
        raise DataTableError(str(exc)) from exc


def sanitize_filename_stem(title: str) -> str:
    text = _BAD_NAME.sub("_", (title or "").strip())
    text = text.strip(" .")
    return text or "未命名数据表"


def unique_xlsx_path(folder: Path, stem: str) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    candidate = folder / f"{stem}.xlsx"
    if not candidate.exists():
        return candidate
    n = 2
    while True:
        candidate = folder / f"{stem}-{n}.xlsx"
        if not candidate.exists():
            return candidate
        n += 1


def create_blank_workbook(
    project_root: Path, leg_name: str, test_name: str, title: str
) -> DataTableRef:
    """Create an empty xlsx in the attachment folder; return an index ref."""
    leg = _require_leg_name(leg_name)
    test = _require_usable_test_name(test_name)
    name = (title or "").strip()
    if not name:
        raise DataTableError("请输入数据表标题")
    folder = attachment_dir(project_root, leg, test)
    stem = sanitize_filename_stem(name)
    dest = unique_xlsx_path(folder, stem)
    wb = Workbook()
    wb.save(dest)
    wb.close()
    rel = dest.relative_to(Path(project_root)).as_posix()
    return DataTableRef(title=name, relative_path=rel)


def upload_existing_xlsx(
    project_root: Path, leg_name: str, test_name: str, source: Path
) -> DataTableRef:
    """Copy an existing .xlsx into the attachment folder; title is the filename stem."""
    leg = _require_leg_name(leg_name)
    test = _require_usable_test_name(test_name)
    src = Path(source)
    if not src.is_file():
        raise DataTableError("找不到所选 Excel 文件")
    if src.suffix.lower() != ".xlsx":
        raise DataTableError("仅支持 .xlsx 文件")
    folder = attachment_dir(project_root, leg, test)
    stem = sanitize_filename_stem(src.stem)
    dest = unique_xlsx_path(folder, stem)
    shutil.copy2(src, dest)
    title = stem
    rel = dest.relative_to(Path(project_root)).as_posix()
    return DataTableRef(title=title, relative_path=rel)


def list_data_table_templates(templates_dir: Path | None = None) -> List[Path]:
    """List .xlsx files in the app-level templates folder (sorted by name)."""
    folder = Path(templates_dir) if templates_dir is not None else default_templates_dir()
    if not folder.is_dir():
        return []
    return sorted(
        (p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"),
        key=lambda p: p.name.lower(),
    )


def copy_from_template(
    project_root: Path, leg_name: str, test_name: str, template_path: Path
) -> DataTableRef:
    """Copy a template .xlsx into the attachment folder (same semantics as upload)."""
    return upload_existing_xlsx(project_root, leg_name, test_name, template_path)


def _col1_has_content(ws) -> bool:
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if _is_nonempty(cell.value):
                return True
    return False


def _sample_id_col_merge_bottom(ws) -> dict[int, int]:
    """Row index in column A -> bottom row of its vertical merge (or itself)."""
    out: dict[int, int] = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_col > 1 or rng.max_col < 1:
            continue
        for r in range(rng.min_row, rng.max_row + 1):
            out[r] = max(out.get(r, r), rng.max_row)
    return out


def _sample_id_start_row(ws) -> int:
    """First row below existing sheet content; empty sheet starts at row 2."""
    bbox = _used_bbox(ws)
    if bbox is None:
        return 2
    default_start = bbox[2] + 1
    merge_bottom = _sample_id_col_merge_bottom(ws)
    header_end = max(1, default_start - 1)
    for r in range(1, header_end + 1):
        val = ws.cell(row=r, column=1).value
        if is_sample_id_column_key(str(val or "").strip()):
            return max(default_start, merge_bottom.get(r, r) + 1)
    return default_start


def _sample_id_column_header_text() -> str:
    return table_header_label("样品编号", "Sample No.", "中英文")


def _col1_has_sample_id_header(ws, header_end_row: int) -> bool:
    for r in range(1, header_end_row + 1):
        if is_sample_id_column_key(str(ws.cell(row=r, column=1).value or "").strip()):
            return True
    return False


def _write_sample_id_column_header(ws, header_end_row: int) -> None:
    """Write bilingual 样品编号 / Sample No. into column A for the header band."""
    if header_end_row < 1 or _col1_has_sample_id_header(ws, header_end_row):
        return
    ws.cell(row=1, column=1, value=_sample_id_column_header_text())
    if header_end_row > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=header_end_row, end_column=1)


def _should_insert_sample_id_column(ws, start_row: int) -> bool:
    """Left-insert only when col 1 holds non-sample-id content."""
    header_end_row = max(1, start_row - 1)
    if _col1_has_sample_id_header(ws, header_end_row):
        return False
    return _col1_has_content(ws)


def import_sample_ids(path: Path, sample_ids: Sequence[str]) -> None:
    """Write sample ids into the first sheet below existing content; insert col 1 if needed."""
    xlsx = Path(path)
    if not xlsx.is_file():
        raise DataTableError("数据表文件不存在")
    ids = [str(s).strip() for s in sample_ids if str(s).strip()]
    wb = load_workbook(xlsx)
    try:
        ws = wb.worksheets[0]
        start_row = _sample_id_start_row(ws)
        if _should_insert_sample_id_column(ws, start_row):
            ws.insert_cols(1)
            start_row = _sample_id_start_row(ws)
        header_end_row = max(1, start_row - 1)
        _write_sample_id_column_header(ws, header_end_row)
        for i, sid in enumerate(ids):
            ws.cell(row=start_row + i, column=1, value=sid)
        wb.save(xlsx)
    finally:
        wb.close()


def _default_app_exists(app_name: str) -> bool:
    """True if a macOS .app or a PATH executable named app_name is present."""
    if sys.platform == "darwin":
        candidates = [
            Path("/Applications") / f"{app_name}.app",
            Path.home() / "Applications" / f"{app_name}.app",
        ]
        if any(c.is_dir() for c in candidates):
            return True
    return shutil.which(app_name) is not None


def resolve_open_argv(
    path: Path,
    *,
    platform: Optional[str] = None,
    app_exists: Optional[Callable[[str], bool]] = None,
) -> List[str]:
    """Build argv to open an xlsx: Excel → WPS → system default. Does not launch."""
    target = str(Path(path))
    plat = platform if platform is not None else sys.platform
    exists = app_exists or _default_app_exists

    if plat == "darwin":
        for name in _EXCEL_APP_NAMES:
            if exists(name):
                return ["open", "-a", name, target]
        for name in _WPS_APP_NAMES:
            if exists(name):
                return ["open", "-a", name, target]
        return ["open", target]

    if plat.startswith("win"):
        for name in ("EXCEL.EXE", "excel"):
            exe = shutil.which(name) if app_exists is None else (name if exists(name) else None)
            if exe:
                return [exe, target]
        for name in ("wps", "wpsoffice"):
            exe = shutil.which(name) if app_exists is None else (name if exists(name) else None)
            if exe:
                return [exe, target]
        return ["cmd", "/c", "start", "", target]

    # Linux / other: PATH then xdg-open
    for name in ("excel", "soffice", "wps", "wpsoffice"):
        if exists(name):
            exe = shutil.which(name) or name
            return [exe, target]
    return ["xdg-open", target]


def open_attachment(
    path: Path,
    *,
    runner: Optional[Callable[..., Any]] = None,
    resolve_argv: Optional[Callable[[Path], List[str]]] = None,
) -> None:
    """Open the attachment in an external spreadsheet app. Raises DataTableError on failure."""
    xlsx = Path(path)
    if not xlsx.is_file():
        raise DataTableError("数据表文件不存在")
    argv = (resolve_argv or resolve_open_argv)(xlsx)
    run = runner or subprocess.run
    try:
        result = run(argv, check=False, capture_output=True, text=True)
    except OSError as exc:
        raise DataTableError(f"无法打开数据表：{exc}") from exc
    code = getattr(result, "returncode", 0)
    if code:
        err = (getattr(result, "stderr", None) or getattr(result, "stdout", None) or "").strip()
        detail = f"：{err}" if err else ""
        raise DataTableError(f"无法打开数据表{detail}")


def _cell_display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:  # NaN
        return ""
    if isinstance(value, str):
        return value
    text = str(value).strip()
    if text in {"", "nan", "NaT", "None"}:
        return ""
    return str(value)


def _is_nonempty(value: Any) -> bool:
    return _cell_display(value) != ""


def _used_bbox(ws) -> Tuple[int, int, int, int] | None:
    """Return 1-based (min_row, min_col, max_row, max_col) for nonempty cells, or None."""
    min_r = min_c = max_r = max_c = None
    for row in ws.iter_rows():
        for cell in row:
            if not _is_nonempty(cell.value):
                continue
            r, c = cell.row, cell.column
            min_r = r if min_r is None else min(min_r, r)
            max_r = r if max_r is None else max(max_r, r)
            min_c = c if min_c is None else min(min_c, c)
            max_c = c if max_c is None else max(max_c, c)
    if min_r is None:
        return None
    return min_r, min_c, max_r, max_c


def _merges_in_bbox(
    merges: Sequence, min_r: int, min_c: int, max_r: int, max_c: int
) -> List[str]:
    out: List[str] = []
    for rng in merges:
        # CellRange has min_row/min_col/max_row/max_col
        if rng.max_row < min_r or rng.min_row > max_r:
            continue
        if rng.max_col < min_c or rng.min_col > max_c:
            continue
        out.append(str(rng))
    return out


def _header_count_from_merges(snap: PreviewSnapshot) -> int:
    """Rows touched by a merge starting at/including the first sheet row."""
    if not snap.values:
        return 0
    origin_r = snap.origin_row or 1
    max_header = 0
    touched = False
    for merge in snap.merges or []:
        try:
            min_c, min_r, max_c, max_r = range_boundaries(merge)
        except Exception:
            continue
        r0 = min_r - origin_r
        r1 = max_r - origin_r
        if r0 <= 0 <= r1 or r0 == 0:
            touched = True
            max_header = max(max_header, r1)
    if touched:
        return max_header + 1
    return 0


def _merge_bottom_by_cell(snap: PreviewSnapshot) -> dict[tuple[int, int], int]:
    """Grid (r,c) -> bottom row index of its merge (or r if unmerged)."""
    origin_r = snap.origin_row or 1
    origin_c = snap.origin_col or 1
    out: dict[tuple[int, int], int] = {}
    for merge in snap.merges or []:
        try:
            min_c, min_r, max_c, max_r = range_boundaries(merge)
        except Exception:
            continue
        r0 = min_r - origin_r
        r1 = max_r - origin_r
        c0 = min_c - origin_c
        c1 = max_c - origin_c
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                out[(r, c)] = max(out.get((r, c), r), r1)
    return out


def _header_count_from_sample_id(snap: PreviewSnapshot) -> int | None:
    """First data row is below 样品编号 label band, or first non-label value in column A."""
    values = snap.values or []
    if not values:
        return None
    merge_bottom = _merge_bottom_by_cell(snap)

    label_data_start: int | None = None
    for r, row in enumerate(values):
        for c, val in enumerate(row):
            if is_sample_id_column_key((val or "").strip()):
                bottom = merge_bottom.get((r, c), r)
                start = bottom + 1
                if label_data_start is None or start > label_data_start:
                    label_data_start = start

    if label_data_start is not None:
        return label_data_start

    for r, row in enumerate(values):
        val = (row[0] if row else "").strip()
        if not val or is_sample_id_column_key(val):
            continue
        return r
    return None


def infer_header_row_count(snap: PreviewSnapshot) -> int:
    """How many top rows repeat as table headers on page breaks in Word export."""
    if not snap.values:
        return 0
    merge_count = _header_count_from_merges(snap)
    sample_count = _header_count_from_sample_id(snap)
    if sample_count is not None:
        return max(merge_count, sample_count, 1)
    return max(merge_count, 1)


# Display-string number: optional sign, digits, optional fractional part (no sci notation).
_NUMERIC_DISPLAY_RE = re.compile(r"^[+-]?(?:\d+\.?\d*|\.\d+)$")


def decimal_places(text: str) -> int | None:
    """Decimal digits after '.' in a display string, or None if not a plain number."""
    s = (text or "").strip()
    if not s or not _NUMERIC_DISPLAY_RE.match(s):
        return None
    if "." in s:
        return len(s.split(".", 1)[1])
    return 0


def parse_numeric_display(text: str) -> float | None:
    """Parse a plain numeric display string; None if not parseable as a number."""
    s = (text or "").strip()
    if not s or not _NUMERIC_DISPLAY_RE.match(s):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _data_region(snap: PreviewSnapshot) -> tuple[int, int, int]:
    """Return (header_rows, nrows, ncols) for the preview grid."""
    values = snap.values or []
    nrows = len(values)
    ncols = max((len(r) for r in values), default=0)
    header_rows = infer_header_row_count(snap) if values else 0
    return header_rows, nrows, ncols


def find_decimal_inconsistencies(
    snap: PreviewSnapshot, *, sample_col: int = 0
) -> List[Tuple[int, int]]:
    """Cells (r,c) whose decimal places differ from the column mode.

    Skips sample column, header rows, empty/non-numeric cells. Compares preview
    display strings as-is (text-filled Excel cells).
    """
    values = snap.values or []
    header_rows, nrows, ncols = _data_region(snap)
    flagged: List[Tuple[int, int]] = []
    for c in range(ncols):
        if c == sample_col:
            continue
        places_by_row: List[Tuple[int, int]] = []
        for r in range(header_rows, nrows):
            row = values[r] if r < len(values) else []
            text = row[c] if c < len(row) else ""
            places = decimal_places(text)
            if places is None:
                continue
            places_by_row.append((r, places))
        if len(places_by_row) < 2:
            continue
        counts: dict[int, int] = {}
        for _, p in places_by_row:
            counts[p] = counts.get(p, 0) + 1
        if len(counts) < 2:
            continue
        mode = max(counts.items(), key=lambda kv: (kv[1], -kv[0]))[0]
        for r, p in places_by_row:
            if p != mode:
                flagged.append((r, c))
    return flagged


def find_out_of_range(
    snap: PreviewSnapshot,
    lo: float,
    hi: float,
    *,
    col: int | None = None,
    sample_col: int = 0,
) -> List[Tuple[int, int]]:
    """Cells (r,c) whose numeric display value is outside [lo, hi].

    Skips sample column, header rows, empty/non-numeric. If ``col`` is set, only
    that column is checked (still skips sample_col).
    """
    if lo > hi:
        lo, hi = hi, lo
    values = snap.values or []
    header_rows, nrows, ncols = _data_region(snap)
    cols = range(ncols) if col is None else [col]
    flagged: List[Tuple[int, int]] = []
    for c in cols:
        if c == sample_col or c < 0 or c >= ncols:
            continue
        for r in range(header_rows, nrows):
            row = values[r] if r < len(values) else []
            text = row[c] if c < len(row) else ""
            val = parse_numeric_display(text)
            if val is None:
                continue
            if val < lo or val > hi:
                flagged.append((r, c))
    return flagged


def read_preview_snapshot(path: Path) -> PreviewSnapshot:
    """Read first sheet as a nonempty bounding box; keep holes; list intersecting merges."""
    xlsx = Path(path)
    if not xlsx.is_file():
        raise DataTableError("数据表文件不存在")
    wb = load_workbook(xlsx, data_only=False)
    try:
        ws = wb.worksheets[0]
        sheet_name = ws.title or ""
        bbox = _used_bbox(ws)
        if bbox is None:
            return PreviewSnapshot(sheet_name=sheet_name, values=[], merges=[])
        min_r, min_c, max_r, max_c = bbox
        values: List[List[str]] = []
        for r in range(min_r, max_r + 1):
            row_vals: List[str] = []
            for c in range(min_c, max_c + 1):
                row_vals.append(_cell_display(ws.cell(r, c).value))
            values.append(row_vals)
        merges = _merges_in_bbox(ws.merged_cells.ranges, min_r, min_c, max_r, max_c)
        return PreviewSnapshot(
            sheet_name=sheet_name,
            values=values,
            merges=merges,
            origin_row=min_r,
            origin_col=min_c,
        )
    finally:
        wb.close()


def resolve_attachment_path(project_root: Path, ref: DataTableRef) -> Path:
    return Path(project_root) / ref.relative_path


def delete_attachment(path: Path) -> None:
    """Remove the xlsx from disk if present. Missing file is a no-op."""
    target = Path(path)
    if target.is_file():
        target.unlink()
