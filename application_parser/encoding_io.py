"""Safe text / Excel I/O for Chinese (UTF-8 / GBK) on Windows and Linux.

精简自 ai_report 的 encoding_io：去掉 Word/docx 依赖，仅保留申请单 Excel 所需。
"""

from __future__ import annotations

import io
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Union

import openpyxl

TEXT_ENCODINGS: tuple[str, ...] = ("utf-8", "utf-8-sig", "gb18030", "gbk", "cp936")
DEFAULT_TEXT_ENCODING = "utf-8"
DEFAULT_TEXT_ERRORS = "replace"


def normalize_unicode_text(text: str) -> str:
    if not text:
        return ""
    return unicodedata.normalize("NFC", text)


def decode_text_bytes(
    data: bytes,
    *,
    encodings: Iterable[str] | None = None,
    errors: str = DEFAULT_TEXT_ERRORS,
) -> str:
    if not data:
        return ""
    for enc in encodings or TEXT_ENCODINGS:
        try:
            return normalize_unicode_text(data.decode(enc))
        except (UnicodeDecodeError, LookupError):
            continue
    return normalize_unicode_text(data.decode(DEFAULT_TEXT_ENCODING, errors=errors))


def safe_text(value: Any) -> str:
    """Coerce Excel cell values to a normalized Unicode str."""
    if value is None:
        return ""
    if isinstance(value, bytes):
        return decode_text_bytes(value)
    if isinstance(value, str):
        return normalize_unicode_text(value.strip())
    if isinstance(value, datetime):
        if (
            value.hour == 0
            and value.minute == 0
            and value.second == 0
            and value.microsecond == 0
        ):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return normalize_unicode_text(str(value).strip())


def _fix_latin1_mojibake(name: str, target_encoding: str) -> str:
    try:
        return name.encode("latin-1").decode(target_encoding)
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name


def normalize_upload_filename(filename: str | None) -> str:
    """Repair multipart filenames mis-decoded as Latin-1 on Windows."""
    if not filename:
        return ""
    name = filename.strip().replace("\\", "/").split("/")[-1]
    if not name:
        return ""
    if any("\u4e00" <= ch <= "\u9fff" for ch in name):
        return name
    for enc in ("utf-8", "gbk", "gb18030"):
        fixed = _fix_latin1_mojibake(name, enc)
        if fixed != name and any("\u4e00" <= ch <= "\u9fff" for ch in fixed):
            return fixed
    return name


def read_binary_file(path: Union[str, Path]) -> bytes:
    with open(path, "rb") as f:
        return f.read()


def load_workbook_from_bytes(file_bytes: bytes, *, data_only: bool = True) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=data_only)


def load_workbook_from_path(path: Union[str, Path], *, data_only: bool = True) -> openpyxl.Workbook:
    return load_workbook_from_bytes(read_binary_file(path), data_only=data_only)
