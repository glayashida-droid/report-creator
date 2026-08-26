"""Sample-info English labels and Sheet2 CN/EN row merge."""

from openpyxl import Workbook

from application_parser.excel_parser import parse_application_sheet2
from application_parser.report_language import english_sample_field_to_cn
from src.language_copy import field_label


def test_english_sample_field_to_cn_common_rows():
    assert english_sample_field_to_cn("Color") == "颜色"
    assert english_sample_field_to_cn("Colour") == "颜色"
    assert english_sample_field_to_cn("Material Code") == "材料编号"
    assert english_sample_field_to_cn("Material No.") == "材料编号"
    assert english_sample_field_to_cn("Material Trademark") == "材料牌号"
    assert english_sample_field_to_cn("Production Date") == "生产日期"
    assert english_sample_field_to_cn("Sample Batch") == "样品批号"
    # Ambiguous bare Material: merge into previous Chinese row, not 材料牌号
    assert english_sample_field_to_cn("Material") is None


def test_overview_english_labels_for_extra_sample_fields():
    assert field_label("颜色", "英文") == "Color"
    assert field_label("材料编号", "英文") == "Material Code"
    assert field_label("材质", "英文") == "Material"
    assert field_label("生产日期", "英文") == "Production Date"
    assert field_label("样品批号", "英文") == "Sample Batch"


def test_sheet2_merges_material_and_color_english_rows():
    wb = Workbook()
    sheet = wb.active
    rows = [
        ("颜色", "黑色"),
        ("Color", "Black"),
        ("材料编号", "/"),
        ("Material Code", "/"),
        ("材质", "/"),
        ("Material", "/"),
        ("生产日期", "2026-07-10"),
        ("Production Date", "2026-07-10"),
        ("样品批号", "46213"),
        ("Sample Batch", "46213"),
        ("车型", "VW316-6"),
        ("Model", "VW316-6"),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        sheet.cell(i, 1, label)
        sheet.cell(i, 2, value)

    sample_info, candidates = parse_application_sheet2(sheet)

    assert "材料牌号" not in sample_info
    assert set(sample_info) >= {
        "颜色",
        "材料编号",
        "材质",
        "生产日期",
        "样品批号",
        "车型",
    }
    assert sample_info["颜色"] == "黑色"
    assert "Black" in candidates["颜色"]
    assert sample_info["生产日期"] == "2026-07-10"
    assert sample_info["样品批号"] == "46213"
    assert sample_info["车型"] == "VW316-6"
