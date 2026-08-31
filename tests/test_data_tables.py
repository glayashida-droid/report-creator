import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.io.data_tables import (
    DataTableError,
    PreviewSnapshot,
    attachment_dir,
    copy_from_template,
    create_blank_workbook,
    decimal_places,
    delete_attachment,
    find_decimal_inconsistencies,
    find_out_of_range,
    import_sample_ids,
    infer_header_row_count,
    list_data_table_templates,
    open_attachment,
    parse_numeric_display,
    read_preview_snapshot,
    resolve_open_argv,
    retarget_node_data_tables,
    rewrite_test_dir_in_relative_path,
    upload_existing_xlsx,
)
from src.io.test_photos import test_dir_key as leg_test_dir_key
from src.models.project_state import DataTableRef, ProjectState, TestLeg, TestNode

LEG = "Leg 1"


def test_create_blank_workbook_writes_xlsx_and_returns_ref():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        ref = create_blank_workbook(root, LEG, "高温试验", "工况记录")
        assert ref.title == "工况记录"
        path = root / ref.relative_path
        assert path.is_file()
        assert path.suffix.lower() == ".xlsx"
        assert path.parent == attachment_dir(root, LEG, "高温试验")
        assert ref.relative_path == (
            f"3.测试组/{leg_test_dir_key(LEG, '高温试验')}/数据表附件/工况记录.xlsx"
        )
        wb = load_workbook(path)
        assert wb.sheetnames
        wb.close()


def test_create_blank_rejects_unusable_test_name():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        try:
            create_blank_workbook(root, LEG, "请选择试验...", "表")
            raise AssertionError("expected DataTableError")
        except DataTableError:
            pass
        assert not (root / "3.测试组").exists()


def test_create_blank_unique_filename_does_not_overwrite():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        first = create_blank_workbook(root, LEG, "高温试验", "表A")
        second = create_blank_workbook(root, LEG, "高温试验", "表A")
        assert first.relative_path != second.relative_path
        assert (root / first.relative_path).is_file()
        assert (root / second.relative_path).is_file()


def test_data_table_index_round_trips_in_project_json():
    state = ProjectState(project_id="P1")
    node = TestNode(
        test_name="高温试验",
        data_tables=[
            DataTableRef(
                title="工况记录",
                relative_path="3.测试组/Leg 1-高温试验/数据表附件/工况记录.xlsx",
            )
        ],
    )
    state.legs.append(TestLeg(leg_id="L1", leg_name="Leg 1", nodes=[node]))
    path = Path(".scratch/test_data_tables_state.json")
    state.save_to_file(str(path))
    try:
        loaded = ProjectState.load_from_file(str(path))
        refs = loaded.legs[0].nodes[0].data_tables
        assert len(refs) == 1
        assert refs[0].title == "工况记录"
        assert refs[0].relative_path == "3.测试组/Leg 1-高温试验/数据表附件/工况记录.xlsx"
    finally:
        if path.exists():
            path.unlink()


def _write_bbox_fixture(path: Path) -> None:
    """3x2 used area with a hole at B2 and a merge on row 3."""
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws["A1"] = "列甲"
    ws["B1"] = "列乙"
    ws["C1"] = "列丙"
    ws["A2"] = 10
    # B2 left empty on purpose
    ws["C2"] = 30
    ws.merge_cells("A3:B3")
    ws["A3"] = "合并区"
    wb.save(path)
    wb.close()


def test_upload_copies_xlsx_and_titles_from_filename():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        src = Path(tmp) / "outside" / "工况记录表.xlsx"
        src.parent.mkdir()
        _write_bbox_fixture(src)
        ref = upload_existing_xlsx(root, LEG, "高温试验", src)
        assert ref.title == "工况记录表"
        dest = root / ref.relative_path
        assert dest.is_file()
        assert dest.parent == attachment_dir(root, LEG, "高温试验")
        assert dest.name == "工况记录表.xlsx"
        assert dest.read_bytes() == src.read_bytes()


def test_upload_same_name_gets_suffix_not_overwrite():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        src = Path(tmp) / "outside" / "同名.xlsx"
        src.parent.mkdir()
        _write_bbox_fixture(src)
        first = upload_existing_xlsx(root, LEG, "高温试验", src)
        second = upload_existing_xlsx(root, LEG, "高温试验", src)
        assert first.title == "同名"
        assert second.title == "同名"
        assert first.relative_path != second.relative_path
        assert (root / first.relative_path).is_file()
        assert (root / second.relative_path).is_file()


def test_upload_then_preview_snapshot_keeps_bbox_empties():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        src = Path(tmp) / "outside" / "夹具.xlsx"
        src.parent.mkdir()
        _write_bbox_fixture(src)
        ref = upload_existing_xlsx(root, LEG, "高温试验", src)
        snap = read_preview_snapshot(root / ref.relative_path)
        assert snap.values == [
            ["列甲", "列乙", "列丙"],
            ["10", "", "30"],
            ["合并区", "", ""],
        ]
        assert "A3:B3" in snap.merges


def test_preview_snapshot_is_bbox_keeps_empty_cells_and_merges():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "fixture.xlsx"
        _write_bbox_fixture(path)
        snap = read_preview_snapshot(path)
        assert snap.sheet_name == "数据"
        assert snap.values == [
            ["列甲", "列乙", "列丙"],
            ["10", "", "30"],
            ["合并区", "", ""],
        ]
        assert "A3:B3" in snap.merges


def test_preview_refresh_rereads_disk_changes():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "live.xlsx"
        _write_bbox_fixture(path)
        before = read_preview_snapshot(path)
        assert before.values[1][2] == "30"

        wb = load_workbook(path)
        ws = wb.active
        ws["C2"] = 99
        ws["D1"] = "列丁"
        wb.save(path)
        wb.close()

        after = read_preview_snapshot(path)
        assert after.values[0] == ["列甲", "列乙", "列丙", "列丁"]
        assert after.values[1][2] == "99"
        assert after.values != before.values


def test_import_sample_ids_empty_col1_writes_from_row2():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "empty.xlsx"
        wb = Workbook()
        wb.save(path)
        wb.close()
        import_sample_ids(path, ["A01", "A02", "A03"])
        wb = load_workbook(path)
        ws = wb.active
        assert ws["A1"].value == "样品编号\nSample No."
        assert ws["A2"].value == "A01"
        assert ws["A3"].value == "A02"
        assert ws["A4"].value == "A03"
        wb.close()


def test_import_sample_ids_inserts_col_when_col1_has_content():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "filled.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "结果"
        ws["B1"] = "备注"
        wb.save(path)
        wb.close()
        import_sample_ids(path, ["S1", "S2"])
        wb = load_workbook(path)
        ws = wb.active
        assert ws["A1"].value == "样品编号\nSample No."
        assert ws["A2"].value == "S1"
        assert ws["A3"].value == "S2"
        assert ws["B1"].value == "结果"
        assert ws["C1"].value == "备注"
        wb.close()


def test_import_sample_ids_starts_below_multi_row_header():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "two_header.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("B1:D1")
        ws["B1"] = "试验后 after test"
        ws["B2"] = "桥路电阻"
        ws["C2"] = "短路电阻"
        ws["D2"] = "绝缘电阻"
        wb.save(path)
        wb.close()
        import_sample_ids(path, ["A01", "A02", "A03"])
        wb = load_workbook(path)
        ws = wb.active
        assert ws["A1"].value == "样品编号\nSample No."
        assert ws["A3"].value == "A01"
        assert ws["A4"].value == "A02"
        assert ws["A5"].value == "A03"
        assert ws["B1"].value == "试验后 after test"
        assert ws["B2"].value == "桥路电阻"
        wb.close()


def test_import_sample_ids_skips_header_when_already_present():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "labeled.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "样品编号 Sample No."
        ws.merge_cells("A1:A2")
        ws["B1"] = "值"
        wb.save(path)
        wb.close()
        import_sample_ids(path, ["X1"])
        wb = load_workbook(path)
        ws = wb.active
        assert ws["A1"].value == "样品编号 Sample No."
        assert ws["A3"].value == "X1"
        wb.close()


def test_infer_header_row_count_two_row_header_without_vertical_merge():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "two_header.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("B1:D1")
        ws["B1"] = "试验后 after test"
        ws["B2"] = "桥路电阻"
        ws["C2"] = "短路电阻"
        ws["D2"] = "绝缘电阻"
        ws["A3"] = "A22607480801-A01"
        wb.save(path)
        wb.close()
        snap = read_preview_snapshot(path)
        assert infer_header_row_count(snap) == 2


def test_infer_header_row_count_sample_no_label_with_merge_band():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "merge_header.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "样品编号 Sample No."
        ws.merge_cells("A1:A4")
        ws["B1"] = "试验前"
        ws["E1"] = "试验后"
        ws.merge_cells("B1:D1")
        ws.merge_cells("E1:G1")
        ws["B2"] = "Before test"
        ws["E2"] = "After test"
        ws.merge_cells("B2:D2")
        ws.merge_cells("E2:G2")
        ws["B3"] = "桥路电阻"
        ws["C3"] = "短路电阻"
        ws["D3"] = "绝缘电阻"
        ws["E3"] = "桥路电阻"
        ws["F3"] = "短路电阻"
        ws["G3"] = "绝缘电阻"
        ws["B4"] = "(Ω)"
        ws["C4"] = "(Ω)"
        ws["D4"] = "(MΩ)"
        ws["E4"] = "(Ω)"
        ws["F4"] = "(Ω)"
        ws["G4"] = "(MΩ)"
        ws["A5"] = "TP-1"
        wb.save(path)
        wb.close()
        snap = read_preview_snapshot(path)
        assert infer_header_row_count(snap) == 4


def test_import_sample_ids_missing_file_raises():
    try:
        import_sample_ids(Path("/tmp/no-such-data-table.xlsx"), ["A01"])
        raise AssertionError("expected DataTableError")
    except DataTableError:
        pass


def test_resolve_open_argv_prefers_excel_then_wps_then_default():
    excel = resolve_open_argv(
        Path("/tmp/a.xlsx"),
        platform="darwin",
        app_exists=lambda name: name == "Microsoft Excel",
    )
    assert excel == ["open", "-a", "Microsoft Excel", "/tmp/a.xlsx"]

    wps = resolve_open_argv(
        Path("/tmp/a.xlsx"),
        platform="darwin",
        app_exists=lambda name: name == "wpsoffice",
    )
    assert wps == ["open", "-a", "wpsoffice", "/tmp/a.xlsx"]

    fallback = resolve_open_argv(
        Path("/tmp/a.xlsx"),
        platform="darwin",
        app_exists=lambda _name: False,
    )
    assert fallback == ["open", "/tmp/a.xlsx"]


def test_open_attachment_uses_injected_runner_and_raises_on_failure():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "a.xlsx"
        wb = Workbook()
        wb.save(path)
        wb.close()

        calls = []

        def runner(argv, **_kwargs):
            calls.append(argv)

            class R:
                returncode = 0

            return R()

        open_attachment(
            path,
            runner=runner,
            resolve_argv=lambda p: ["open", "-a", "Microsoft Excel", str(p)],
        )
        assert calls == [["open", "-a", "Microsoft Excel", str(path)]]

        def fail_runner(_argv, **_kwargs):
            class R:
                returncode = 1
                stderr = "boom"

            return R()

        try:
            open_attachment(
                path,
                runner=fail_runner,
                resolve_argv=lambda p: ["open", str(p)],
            )
            raise AssertionError("expected DataTableError")
        except DataTableError as exc:
            assert "无法打开" in str(exc) or "打开" in str(exc)


def test_list_templates_empty_or_missing_is_safe():
    with tempfile.TemporaryDirectory() as tmp:
        missing = Path(tmp) / "nope"
        assert list_data_table_templates(missing) == []
        empty = Path(tmp) / "empty"
        empty.mkdir()
        assert list_data_table_templates(empty) == []


def test_copy_from_template_copies_and_titles_from_filename():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp) / "project"
        root.mkdir()
        templates = Path(tmp) / "templates"
        templates.mkdir()
        src = templates / "高温记录.xlsx"
        _write_bbox_fixture(src)
        ref = copy_from_template(root, LEG, "高温试验", src)
        assert ref.title == "高温记录"
        dest = root / ref.relative_path
        assert dest.is_file()
        assert dest.parent == attachment_dir(root, LEG, "高温试验")
        assert dest.read_bytes() == src.read_bytes()


def test_list_templates_returns_xlsx_sorted_by_name():
    with tempfile.TemporaryDirectory() as tmp:
        folder = Path(tmp)
        (folder / "b.xlsx").write_bytes(b"PK")
        (folder / "a.xlsx").write_bytes(b"PK")
        (folder / "readme.txt").write_text("x")
        (folder / "skip").mkdir()
        names = [p.name for p in list_data_table_templates(folder)]
        assert names == ["a.xlsx", "b.xlsx"]


def test_delete_attachment_removes_file_missing_is_noop():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        ref = create_blank_workbook(root, LEG, "高温试验", "待删")
        path = root / ref.relative_path
        assert path.is_file()
        delete_attachment(path)
        assert not path.exists()
        delete_attachment(path)  # missing → no error


def test_rewrite_test_dir_in_relative_path():
    old_key = leg_test_dir_key(LEG, "湿热循环")
    new_key = leg_test_dir_key(LEG, "前湿热循环")
    assert (
        rewrite_test_dir_in_relative_path(
            f"3.测试组/{old_key}/数据表附件/工况.xlsx", old_key, new_key
        )
        == f"3.测试组/{new_key}/数据表附件/工况.xlsx"
    )
    assert (
        rewrite_test_dir_in_relative_path(
            "3.测试组/Leg 2-其他/数据表附件/a.xlsx", old_key, new_key
        )
        == "3.测试组/Leg 2-其他/数据表附件/a.xlsx"
    )


def test_retarget_node_data_tables():
    old_key = leg_test_dir_key(LEG, "湿热循环")
    new_key = leg_test_dir_key(LEG, "前湿热循环")
    node = TestNode(
        test_name="前湿热循环",
        data_tables=[
            DataTableRef(
                title="工况.xlsx",
                relative_path=f"3.测试组/{old_key}/数据表附件/工况.xlsx",
            )
        ],
    )
    retarget_node_data_tables(node, old_key, new_key)
    assert node.data_tables[0].relative_path == (
        f"3.测试组/{new_key}/数据表附件/工况.xlsx"
    )


def test_decimal_places_from_display_string():
    assert decimal_places("1.20") == 2
    assert decimal_places("1.2") == 1
    assert decimal_places("2") == 0
    assert decimal_places("-1.50") == 2
    assert decimal_places("") is None
    assert decimal_places("A01") is None
    assert decimal_places("1.2Ω") is None


def test_parse_numeric_display():
    assert parse_numeric_display("1.20") == 1.2
    assert parse_numeric_display("-3") == -3.0
    assert parse_numeric_display("样品") is None


def test_find_decimal_inconsistencies_flags_minority_in_column():
    snap = PreviewSnapshot(
        sheet_name="Sheet",
        values=[
            ["样品编号", "桥路", "短路"],
            ["A01", "1.20", "2.5"],
            ["A02", "1.2", "2.50"],
            ["A03", "1.21", "2.5"],
        ],
        merges=[],
    )
    # Header inferred as 1 (sample label in A1); col1 mode is 2 decimals → flag 1.2
    flagged = find_decimal_inconsistencies(snap)
    assert (2, 1) in flagged
    assert (1, 1) not in flagged
    assert (3, 1) not in flagged
    # col2 mode is 1 decimal → flag 2.50
    assert (2, 2) in flagged
    assert (1, 2) not in flagged


def test_find_decimal_inconsistencies_skips_sample_col_and_consistent():
    snap = PreviewSnapshot(
        sheet_name="Sheet",
        values=[
            ["样品编号", "值"],
            ["A01", "1.0"],
            ["A02", "2.0"],
        ],
        merges=[],
    )
    assert find_decimal_inconsistencies(snap) == []


def test_find_out_of_range_whole_table_and_column():
    snap = PreviewSnapshot(
        sheet_name="Sheet",
        values=[
            ["样品编号", "桥路", "短路"],
            ["A01", "1.2", "9.0"],
            ["A02", "5.0", "2.5"],
        ],
        merges=[],
    )
    all_hit = find_out_of_range(snap, 0.0, 3.0)
    assert set(all_hit) == {(1, 2), (2, 1)}
    col_hit = find_out_of_range(snap, 0.0, 3.0, col=1)
    assert col_hit == [(2, 1)]
    assert find_out_of_range(snap, 0.0, 3.0, col=0) == []


if __name__ == "__main__":
    test_create_blank_workbook_writes_xlsx_and_returns_ref()
    test_create_blank_rejects_unusable_test_name()
    test_create_blank_unique_filename_does_not_overwrite()
    test_data_table_index_round_trips_in_project_json()
    test_upload_copies_xlsx_and_titles_from_filename()
    test_upload_same_name_gets_suffix_not_overwrite()
    test_upload_then_preview_snapshot_keeps_bbox_empties()
    test_preview_snapshot_is_bbox_keeps_empty_cells_and_merges()
    test_preview_refresh_rereads_disk_changes()
    test_import_sample_ids_empty_col1_writes_from_row2()
    test_import_sample_ids_inserts_col_when_col1_has_content()
    test_import_sample_ids_starts_below_multi_row_header()
    test_import_sample_ids_missing_file_raises()
    test_resolve_open_argv_prefers_excel_then_wps_then_default()
    test_open_attachment_uses_injected_runner_and_raises_on_failure()
    test_list_templates_empty_or_missing_is_safe()
    test_copy_from_template_copies_and_titles_from_filename()
    test_list_templates_returns_xlsx_sorted_by_name()
    test_delete_attachment_removes_file_missing_is_noop()
    test_retarget_node_data_tables()
    print("test_data_tables: ok")