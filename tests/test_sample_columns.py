"""Multi-column sample info: parse, merge, and project state."""

from pathlib import Path

from openpyxl import Workbook

from application_parser.excel_parser import parse_application, parse_application_sheet2_columns
from application_parser.models import ApplicationData, FileSource
from src.application_ingest import apply_application_data
from src.models.project_state import ProjectState
from src.sample_columns import ALL_SAMPLE_COLUMNS, merge_column_field_values, merge_sample_column_dicts


def _two_column_sheet():
    wb = Workbook()
    sheet = wb.active
    rows = [
        ("样品序号", "001", "002"),
        ("样品名称", "左安全气帘总成", "右安全气帘总成"),
        ("Sample Name", "CURTAIN ASSY LH", "CURTAIN ASSY RH"),
        ("零件号", "6608678545", "6608678544"),
        ("Part No.", "6608678545", "6608678544"),
        ("生产日期", "2026/8/6", "2026/8/6"),
        ("Production Date", "2026/8/6", "2026/8/6"),
        ("买家", "GEELY", "GEELY"),
        ("Buyer", "GEELY", "GEELY"),
    ]
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row):
            sheet.cell(i, j + 1, val)
    return sheet


def test_parse_sheet2_two_columns():
    sheet = _two_column_sheet()
    cols_cn, cols_en, labels = parse_application_sheet2_columns(sheet)

    assert labels == ["001", "002"]
    assert len(cols_cn) == 2
    assert cols_cn[0]["样品名称"] == "左安全气帘总成"
    assert cols_cn[1]["样品名称"] == "右安全气帘总成"
    assert cols_cn[0]["零件号"] == "6608678545"
    assert cols_cn[1]["零件号"] == "6608678544"
    assert cols_en[0]["样品名称"] == "CURTAIN ASSY LH"
    assert cols_en[1]["样品名称"] == "CURTAIN ASSY RH"


def _single_column_sheet_with_styled_empty_trailing_column():
    """模板 C 列仅有样式、无数据时不应多出 002 tab。"""
    wb = Workbook()
    sheet = wb.active
    rows = [
        ("申请单号", "A22600280178"),
        ("样品序号", "001"),
        ("样品名称", "微型执行器"),
        ("Sample Name", "MA"),
        ("零件号", "6NW 011.122-09"),
        ("送样数量", "4"),
    ]
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row):
            sheet.cell(i, j + 1, val)
    # 模拟模板预留列：C 列被 openpyxl 补齐为空单元格
    for i in range(1, len(rows) + 1):
        sheet.cell(i, 3, "")
    return sheet


def test_parse_sheet2_ignores_trailing_empty_column():
    sheet = _single_column_sheet_with_styled_empty_trailing_column()
    cols_cn, cols_en, labels = parse_application_sheet2_columns(sheet)

    assert len(cols_cn) == 1
    assert labels == ["001"]
    assert cols_cn[0]["样品名称"] == "微型执行器"
    assert cols_en[0]["样品名称"] == "MA"


def test_parse_real_single_column_application_file():
    path = Path("data/A2260028017801/1.接样组/A22600280178.xlsx")
    if not path.is_file():
        return
    data = parse_application(path.read_bytes(), path.name)
    assert len(data.sample_columns_cn) == 1
    assert data.sample_column_tab_labels == ["001"]
    assert data.sample_columns_cn[0]["样品名称"] == "微型执行器"


def test_ingest_single_column_does_not_show_multi_tabs():
    data = ApplicationData(
        source=FileSource(file_type="application", filename="a.xlsx"),
        applicant_name="委托方",
        applicant_address="地址",
        applicant_name_cn="委托方",
        sample_columns_cn=[
            {"申请单号": "A22600280178", "样品名称": "微型执行器"},
        ],
        sample_columns_en=[{"样品名称": "MA"}],
        sample_column_tab_labels=["001"],
    )
    state = ProjectState(project_id="A1")
    apply_application_data(state, data)
    assert not state.has_multiple_sample_columns()
    assert state.sample_column_tab_labels == ["001"]


def test_merge_dedupes_same_values():
    assert merge_column_field_values(["2026/8/6", "2026/8/6"]) == "2026/8/6"
    assert merge_column_field_values(["GEELY", "GEELY"]) == "GEELY"
    assert (
        merge_column_field_values(["左安全气帘总成", "右安全气帘总成"])
        == "左安全气帘总成/右安全气帘总成"
    )


def test_merge_sample_column_dicts_all_mode():
    columns = [
        {"样品名称": "左安全气帘总成", "零件号": "6608678545", "生产日期": "2026/8/6", "买家": "GEELY"},
        {"样品名称": "右安全气帘总成", "零件号": "6608678544", "生产日期": "2026/8/6", "买家": "GEELY"},
    ]
    merged = merge_sample_column_dicts(columns)
    assert merged["样品名称"] == "左安全气帘总成/右安全气帘总成"
    assert merged["零件号"] == "6608678545/6608678544"
    assert merged["生产日期"] == "2026/8/6"
    assert merged["买家"] == "GEELY"


def test_apply_application_data_multi_column_and_all_tab():
    data = ApplicationData(
        source=FileSource(file_type="application", filename="a.xlsx"),
        applicant_name="委托方",
        applicant_address="地址",
        applicant_name_cn="委托方",
        sample_columns_cn=[
            {"申请单号": "A22606889781", "样品名称": "左安全气帘总成", "买家": "GEELY"},
            {"申请单号": "A22606889781", "样品名称": "右安全气帘总成", "买家": "GEELY"},
        ],
        sample_columns_en=[
            {"样品名称": "CURTAIN ASSY LH", "买家": "GEELY"},
            {"样品名称": "CURTAIN ASSY RH", "买家": "GEELY"},
        ],
        sample_column_tab_labels=["001", "002"],
        sample_info={"申请单号": "A22606889781", "样品名称": "左安全气帘总成"},
    )
    state = ProjectState(project_id="A1")
    apply_application_data(state, data)

    assert state.has_multiple_sample_columns()
    assert state.application_fields["样品名称"] == "左安全气帘总成"

    state.set_active_sample_column(1)
    assert state.application_fields["样品名称"] == "右安全气帘总成"
    assert state.application_fields_en["样品名称"] == "CURTAIN ASSY RH"

    state.set_active_sample_column(ALL_SAMPLE_COLUMNS)
    assert state.application_fields["样品名称"] == "左安全气帘总成/右安全气帘总成"
    assert state.application_fields_en["样品名称"] == "CURTAIN ASSY LH/CURTAIN ASSY RH"
    assert state.application_fields["买家"] == "GEELY"
    assert dict(state.iter_overview_fields("中文"))["样品名称"] == "左安全气帘总成/右安全气帘总成"
